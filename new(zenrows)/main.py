import time
import re
import openpyxl
from zenrows import ZenRowsClient
from bs4 import BeautifulSoup
import asyncio

zenrows_key = ""

def read_xlsx(file):
    wb = openpyxl.load_workbook(file)
    worksheet = wb.active
    keys = []
    data = []
    for row in worksheet.iter_rows():
        if len(keys):
            line = {}
            i = 0
            for cell in row:
                line[keys[i]] = cell.value
                i += 1
            data.append(line)
        else:
            for cell in row:
                keys.append(cell.value)
    return data

def write_xlsx(file, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, key in enumerate(data[0].keys(), start=1):
        ws.cell(row=1, column=col, value=key)

    for row, item in enumerate(data, start=2):
        for col, value in enumerate(item.values(), start=1):
            ws.cell(row=row, column=col, value=value)
    
    wb.save(file)

def check_entity(name):
    if "CORP" in name or "LLC" in name or "LP" in name or "PC" in name or "CORPORATION" in name or "LLP" in name or "TRUST" in name or "REALTY" in name:
        return 1
    else:
        return 0

def build_url(data, entity_type, owner_num):
    if entity_type:
        address = data["Mailing Address"]
        citystatezip = data["Mailing City St  Zip"]
        search_url = "https://www.truepeoplesearch.com/resultaddress?streetaddress=" + address.replace(" ", "%20") + "&citystatezip=" + citystatezip.replace(" ", "%20")
    else:
        if owner_num == 1:
            name = data["Owner name 01"]
        if owner_num == 2:
            name = data["Owner name 02"]
        citystatezip = data["Mailing City St  Zip"]
        search_url = "https://www.truepeoplesearch.com/resultaddress?name=" + name.replace(" ", "%20") + "&citystatezip=" + citystatezip.replace(" ", "%20")

    return search_url


async def scrap(data, client, params):
    age_regex = re.compile(r'Age \d+ \([A-Za-z]{3} \d+\)')

    first_name = data["Owner name 01"]
    entity_type_1 = check_entity(first_name)

    search_url = build_url(data, entity_type_1, 1)

    print("\n")
    print(first_name)
    print(search_url)

    response = await client.get_async(search_url, params=params)

    soup = BeautifulSoup(response.text, 'html.parser')
    a_tag = soup.find('a', {'class': 'btn btn-success btn-lg detail-link shadow-form'})

    try:
        detail_url = "https://www.truepeoplesearch.com" + a_tag['href']

        print(detail_url)

        detail_response = await client.get_async(detail_url, params=params)

        with open("html.txt", "w", encoding="UTF-8") as output:
            output.write(detail_response.text)

        detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
        
        owner = detail_soup.select_one('h1.oh1')
        if entity_type_1:
            data["Entity Owner's Name (Only if ENTITY)"] = owner.text
        
        try:
            age_span_1 = detail_soup.find('span', string=age_regex)    
            data["Age1"] = age_span_1.text.strip()
        except:
            data["Age1"] = "Age Unknown"
        
        phone_numbers_1 = []
        phone_divs = detail_soup.select('div.row div.col-12')
        for phone_div in phone_divs:
            try:
                phone_span = phone_div.select_one('a[data-link-to-more="phone"] span[itemprop="telephone"]')
                phone = phone_span.text.strip()

                phone_type_span = phone_div.select_one('span.smaller')
                phone_type = phone_type_span.text.strip()

                phone_info = f"{phone} - {phone_type}"
                if phone_info not in phone_numbers_1:
                    phone_numbers_1.append(phone_info)
            except:
                pass
        i = 1
        while len(phone_numbers_1):
            phone = phone_numbers_1.pop(0)
            if i < 7:
                data["Phone1-"+str(i)] = phone
                i += 1

        emails_1 = []
        email_divs = detail_soup.select('.col > div:last-child')
        for email_div in email_divs:
            text = email_div.text.strip()
            if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text):
                if text != "support@truepeoplesearch.com" and text not in emails_1:
                    emails_1.append(text)

        j = 1
        while len(emails_1):
            email = emails_1.pop(0)
            if j < 3:
                data["Email1-"+str(j)] = email
                j += 1
    except:
        print("Cannot Fine Detail Link")
        with open("error.txt", "a", encoding="UTF-8") as error:
            error.write(first_name + "\n")

    second_name = data["Owner name 02"]    
        
    if second_name:
        entity_type_2 = check_entity(second_name)
        if entity_type_1 == 1 and entity_type_2 == 1:
            pass
        else:
            search_url = build_url(data, entity_type_2, 2)

            print("\n\n")
            print(second_name)
            print(search_url)

            response = await client.get_async(search_url, params=params)
            soup = BeautifulSoup(response.text, 'html.parser')

            a_tag = soup.find('a', {'class': 'btn btn-success btn-lg detail-link shadow-form'})

            try:
                detail_url = "https://www.truepeoplesearch.com" + a_tag['href']

                print(detail_url)

                detail_response = await client.get_async(detail_url, params=params)

                detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
            
                try:
                    age_span_2 = detail_soup.find('span', string=age_regex)    
                    data["Age2"] = age_span_2.text.strip()
                except:
                    data["Age2"] = "Age Unknown"

                phone_numbers_2 = []
                phone_divs = detail_soup.select('div.row div.col-12')
                for phone_div in phone_divs:
                    try:
                        phone_span = phone_div.select_one('a[data-link-to-more="phone"] span[itemprop="telephone"]')
                        phone = phone_span.text.strip()

                        phone_type_span = phone_div.select_one('span.smaller')
                        phone_type = phone_type_span.text.strip()

                        phone_info = f"{phone} - {phone_type}"
                        if phone_info not in phone_numbers_2:
                            phone_numbers_2.append(phone_info)
                    except:
                        pass
                i = 1
                while len(phone_numbers_2):
                    phone = phone_numbers_2.pop(0)
                    if i < 7:
                        data["Phone2-"+str(i)] = phone
                        i += 1

                emails_2 = []
                email_divs = detail_soup.select('.col > div:last-child')
                for email_div in email_divs:
                    text = email_div.text.strip()
                    if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text):
                        if text != "support@truepeoplesearch.com" and text not in emails_2:
                            emails_2.append(text)

                j = 1
                while len(emails_2):
                    email = emails_2.pop(0)
                    if j < 3:
                        data["Email2-"+str(j)] = email
                        j += 1
            except:
                print("Cannot Fine Detail Link")
                with open("error.txt", "a", encoding="UTF-8") as error:
                    error.write(second_name + "\n")

    return data

async def main():
    client = ZenRowsClient(zenrows_key, concurrency=10, retries=1)
    params = {"js_render":"true","antibot":"false","premium_proxy":"true"}

    input_file = 'Upwork Test LLCs.xlsx'
    input_data = read_xlsx(input_file)
    
    output_data = []

    input_temp = input_data[:20]

    tasks = [scrap(item, client, params) for item in input_temp]
    output_data = await asyncio.gather(*tasks)
    # for item in input_temp:
    #     start_time = time.time()
    #     data = scrap(item, client, params)
    #     end_time = time.time()
    #     print(end_time - start_time)
    #     output_data.append(data)

    write_xlsx("result.xlsx", output_data)

if __name__ == "__main__":
    start_time = time.time()
    asyncio.run(main())
    end_time = time.time()
    print("total time", end_time - start_time)